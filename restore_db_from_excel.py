import argparse
import os
from pathlib import Path

import openpyxl


def _normalize_project(value: str) -> str:
    v = (value or "").strip()
    if not v:
        return ""
    v_up = v.upper().replace("_", "-")
    v_up = " ".join(v_up.split())
    if v_up in {"PILOT X", "PILOT-X"}:
        return "PILOT-X"
    if v_up == "NOVA":
        return "NOVA"
    return v_up.replace(" ", "-")


def _cell_str(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def load_rows(excel_path: Path, sheet: str | None) -> list[dict]:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {str(h).strip(): i for i, h in enumerate(headers) if h is not None}

    required = [
        "Date",
        "Employee ID",
        "Name",
        "Leave Type",
        "What I did yesterday",
        "What I am doing today",
        "Any blockers",
        "Tag",
    ]
    for r in required:
        if r not in idx:
            raise ValueError(f"Missing required column: {r}")

    has_project = "Project" in idx

    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        date = _cell_str(row[idx["Date"]])
        emp_id = _cell_str(row[idx["Employee ID"]])
        name = _cell_str(row[idx["Name"]])
        leave_type = _cell_str(row[idx["Leave Type"]])
        last_updates = _cell_str(row[idx["What I did yesterday"]])
        current_updates = _cell_str(row[idx["What I am doing today"]])
        impediments = _cell_str(row[idx["Any blockers"]])
        tag = _cell_str(row[idx["Tag"]]).lstrip("#").upper()
        project = _normalize_project(_cell_str(row[idx["Project"]])) if has_project else ""

        if not (date and emp_id and name):
            continue

        out.append(
            {
                "date": date,
                "member": name,
                "leave_type": leave_type or None,
                "last_updates": last_updates or "Worked on assigned tasks.",
                "current_updates": current_updates or "Continuing planned work.",
                "impediments": impediments or "No blockers",
                "tags": [tag] if tag else [],
                "project": project or "NOVA",
            }
        )

    return out


def main() -> int:
    parser = argparse.ArgumentParser(description="Restore pmtool DB from an Excel export.")
    parser.add_argument("--excel", required=True, help="Path to the Excel file to import.")
    parser.add_argument("--sheet", default=None, help="Sheet name (defaults to first sheet).")
    parser.add_argument(
        "--db-url",
        default="sqlite:///pmtool.db",
        help="SQLAlchemy DB URL (default: sqlite:///pmtool.db which Flask stores under ./instance).",
    )
    parser.add_argument("--wipe", action="store_true", help="Delete existing employee_updates before import.")
    args = parser.parse_args()

    excel_path = Path(args.excel).expanduser().resolve()
    if not excel_path.exists():
        raise FileNotFoundError(excel_path)

    os.environ["DATABASE_URL"] = args.db_url

    from backend.app import create_app  # noqa: E402
    from backend.db import db  # noqa: E402
    from backend.models import EmployeeUpdate, User  # noqa: E402

    rows = load_rows(excel_path, args.sheet)
    if not rows:
        raise ValueError("No importable rows found in the Excel file.")

    app = create_app()
    with app.app_context():
        employee_user = User.query.filter_by(role="employee").order_by(User.id.asc()).first()
        if not employee_user:
            raise RuntimeError("No seeded employee user found (users table is empty).")

        existing = EmployeeUpdate.query.count()
        if existing and not args.wipe:
            raise RuntimeError(
                f"DB already has {existing} employee_updates. Re-run with --wipe to replace them."
            )
        if args.wipe and existing:
            EmployeeUpdate.query.delete()
            db.session.commit()

        for r in rows:
            db.session.add(
                EmployeeUpdate(
                    date=r["date"],
                    project=r["project"],
                    member=r["member"],
                    leave_type=r["leave_type"],
                    last_updates=r["last_updates"],
                    current_updates=r["current_updates"],
                    impediments=r["impediments"],
                    tags=r["tags"],
                    created_by_id=employee_user.id,
                )
            )
        db.session.commit()

        inserted = EmployeeUpdate.query.count()
        print(f"Imported {inserted} employee_updates from {excel_path}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
