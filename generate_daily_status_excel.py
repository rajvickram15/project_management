import random
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl


def _parse_date(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        return datetime.strptime(value.strip(), "%Y-%m-%d").date()
    raise TypeError(f"Unsupported date value: {value!r}")


def generate(
    source_path: Path,
    output_path: Path,
    start_new: date,
    end_new: date,
    projects: tuple[str, str] = ("Nova", "Pilot X"),
    tags: tuple[str, str] = ("FA", "TCP"),
    records_per_day: int = 20,
    max_total_records: int = 210,
):
    rng = random.Random(42)

    wb = openpyxl.load_workbook(source_path)
    ws = wb[wb.sheetnames[0]]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if "Project" in headers:
        raise ValueError("Source already contains a 'Project' column.")

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows:
        raise ValueError("Source worksheet has no data rows.")

    # Build employee directory from existing rows.
    employees: dict[str, str] = {}
    for r in rows:
        emp_id = r[1]
        name = r[2]
        if emp_id and name and emp_id not in employees:
            employees[str(emp_id)] = str(name)

    # Text templates to keep the generated file realistic.
    yesterday_pool = [
        "I worked on API integration and unit tests",
        "I fixed bugs reported by QA and verified locally",
        "I updated UI components and improved validations",
        "I reviewed PRs and refactored a few modules",
        "I worked on database schema updates and migrations",
        "I investigated an issue and shared findings with the team",
    ]
    today_pool = [
        "I will continue feature development and push changes",
        "I will address review comments and add test coverage",
        "I will work on performance improvements and logging",
        "I will implement remaining endpoints and update docs",
        "I will validate the changes in staging and deploy",
        "I will coordinate with QA and close remaining issues",
    ]
    blockers_pool = [
        "No blockers",
        "Waiting for clarifications on requirements",
        "Pending review from another team",
        "Dependency on API changes from backend team",
        "Need access/credentials for an environment",
    ]

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = ws.title

    out_headers = list(headers) + ["Project"]
    out_ws.append(out_headers)

    def pick_project():
        return projects[0] if rng.random() < 0.5 else projects[1]

    # Copy existing rows with Project populated.
    for r in rows:
        out_ws.append(list(r) + [pick_project()])

    # Generate additional days for all employees for the requested range.
    if start_new > end_new:
        raise ValueError("start_new must be <= end_new")
    if records_per_day <= 0:
        raise ValueError("records_per_day must be positive")
    if max_total_records < len(rows):
        raise ValueError("max_total_records must be >= existing record count")

    employee_items = list(employees.items())

    current = start_new
    while current <= end_new:
        remaining_capacity = max_total_records - (out_ws.max_row - 1)
        if remaining_capacity <= 0:
            break

        day_count = min(records_per_day, len(employee_items), remaining_capacity)
        day_employees = rng.sample(employee_items, k=day_count)

        for emp_id, name in day_employees:
            leave_type = "PL" if rng.random() < 0.08 else "None"
            yesterday = rng.choice(yesterday_pool)
            today = rng.choice(today_pool)
            blockers = rng.choice(blockers_pool)
            tag = tags[0] if rng.random() < 0.35 else tags[1]
            out_ws.append(
                [
                    current.strftime("%Y-%m-%d"),
                    emp_id,
                    name,
                    leave_type,
                    yesterday,
                    today,
                    blockers,
                    tag,
                    pick_project(),
                ]
            )
        current += timedelta(days=1)

    # Basic sizing for readability.
    for col in range(1, out_ws.max_column + 1):
        out_ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 22

    output_path.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(output_path)


if __name__ == "__main__":
    source = Path(r"C:\Users\rkrajvickram\Downloads\simple_daily_status.xlsx")
    out = Path("simple_daily_status_extended_with_project_210.xlsx")

    # Your source already contains dates up to 2026-05-06; add 5 more days (2026-05-07 .. 2026-05-11).
    generate(
        source_path=source,
        output_path=out,
        start_new=date(2026, 5, 7),
        end_new=date(2026, 5, 11),
        projects=("Nova", "Pilot X"),
        tags=("FA", "TCP"),
        records_per_day=20,
        max_total_records=210,
    )
    print(f"Wrote: {out.resolve()}")
